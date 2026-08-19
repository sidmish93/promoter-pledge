"""
Live promoter-pledge screen.

Run:
    python app.py

Then open http://127.0.0.1:5000
"""
from __future__ import annotations

import io
import os
import traceback
from datetime import date

from flask import Flask, jsonify, make_response, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from fetchers import (
    fmt_date,
    last_quarter_end,
    list_invoke_events,
    list_pledged_companies,
    list_promoters_for_companies,
    load_company,
    to_num,
)

app = Flask(__name__)
app.config["TEMPLATES_AUTO_RELOAD"] = True
app.jinja_env.auto_reload = True


@app.route("/")
def index():
    resp = make_response(render_template("index.html"))
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate"
    resp.headers["Pragma"] = "no-cache"
    return resp


@app.route("/api/companies", methods=["POST"])
def companies():
    try:
        rows, meta = list_pledged_companies()
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 502
    return jsonify({
        "companies": rows,
        "count": len(rows),
        "shp_cutoff_label": fmt_date(last_quarter_end()),
        "list_meta": meta,
    })


@app.route("/api/invokes", methods=["POST"])
def invokes():
    try:
        rows, meta = list_invoke_events()
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 502
    return jsonify({
        "events": rows,
        "count": len(rows),
        "meta": meta,
    })


_INVOKE_XLS_COLS = (
    ("event_date_label", "Date"),
    ("company", "Company"),
    ("nse_symbol", "NSE ticker"),
    ("promoter", "Promoter"),
    ("lender", "Lender"),
    ("encumbrance_type", "Encumbrance"),
    ("event_shares", "Shares invoked"),
    ("event_pct_equity", "% of equity"),
    ("pre_event_encumbered_shares", "Pledged before"),
    ("post_event_encumbered_shares", "Pledged after"),
    ("broadcast_label", "Broadcast"),
    ("attachment", "Filing"),
    ("source", "Source"),
)


def _style_sheet(ws, widths=None):
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    ws.freeze_panes = "A2"
    if ws.max_row >= 1 and ws.max_column >= 1:
        ws.auto_filter.ref = ws.dimensions
    ws.row_dimensions[1].height = 32
    if widths:
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w


def _append_rows(ws, cols, rows):
    ws.append([h for _, h in cols])
    for row in rows:
        ws.append([row.get(key) for key, _ in cols])


_COMPANY_XLS_COLS = (
    ("name", "Company"),
    ("nse_symbol", "NSE ticker"),
    ("bse_scripcode", "BSE scrip"),
    ("isin", "ISIN"),
    ("list_exchange", "Exchange tape"),
    ("shp_pledged_shares", "SHP pledged shares"),
    ("shp_pledged_pct_of_promoter", "Pledged % of promoter holding"),
    ("promoter_holding_pct", "Promoter holding %"),
    ("pledged_value_cr", "Pledged value (Rs cr)"),
)

_PROMOTER_XLS_COLS = (
    ("company", "Company"),
    ("nse_symbol", "NSE ticker"),
    ("bse_scripcode", "BSE scrip"),
    ("quarter", "SHP quarter"),
    ("promoter", "Promoter"),
    ("category", "Category"),
    ("holding_shares", "Holding shares"),
    ("holding_pct", "Holding %"),
    ("pledged_shares", "Pledged shares"),
    ("pledged_pct_of_holding", "Pledged % of holding"),
    ("pledged_value_cr", "Pledged value (Rs cr)"),
)


def _companies_meeting_threshold(rows: list, min_cr: float) -> list[dict]:
    out = [r for r in rows if isinstance(r, dict)]
    if min_cr > 0:
        kept = []
        for r in out:
            val = to_num(r.get("pledged_value_cr"))
            if val is not None and val >= min_cr:
                kept.append(r)
        out = kept
    return out[:2000]


def _start_workbook(companies: list[dict], promoters: list[dict], min_cr: float) -> Workbook:
    wb = Workbook()
    notes = wb.active
    notes.title = "How to read"
    notes.append(["Topic", "What it means"])
    notes.append(["Who is in this file", "Only companies that pass the min pledge value on Start."])
    notes.append(["Min pledge (Rs cr)", min_cr if min_cr > 0 else "0 = all"])
    notes.append(["Company rows", len(companies)])
    notes.append(["Promoter rows", len(promoters)])
    notes.append(
        [
            "Promoters sheet",
            "Latest SHP pledged promoters only. Not earlier quarters, not Regulation 31.",
        ]
    )
    notes.append(
        [
            "Missing promoters",
            "A company can appear on Companies with no Promoters row if the SHP table did not load.",
        ]
    )
    _style_sheet(notes, (22, 88))

    ws = wb.create_sheet("Companies")
    _append_rows(ws, _COMPANY_XLS_COLS, companies)
    _style_sheet(ws, (36, 14, 12, 16, 14, 18, 22, 18, 18))

    ps = wb.create_sheet("Promoters")
    _append_rows(ps, _PROMOTER_XLS_COLS, promoters)
    _style_sheet(ps, (36, 14, 12, 16, 32, 16, 16, 12, 16, 18, 18))
    return wb


def _xlsx_response(wb: Workbook, filename: str):
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


def _invoke_workbook(rows: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Invokes"
    headers = [h for _, h in _INVOKE_XLS_COLS]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    for ev in rows:
        ws.append([ev.get(key) for key, _ in _INVOKE_XLS_COLS])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = (14, 36, 14, 32, 32, 14, 16, 12, 16, 16, 14, 40, 10)
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.row_dimensions[1].height = 32
    return wb


@app.route("/api/invokes.xlsx", methods=["POST"])
def invokes_xlsx():
    try:
        data = request.get_json(silent=True) or {}
        rows = data.get("events")
        if not isinstance(rows, list) or not rows:
            rows, _meta = list_invoke_events()
        rows = [r for r in rows if isinstance(r, dict)][:5000]
        if not rows:
            return jsonify({"error": "No invocation filings to download."}), 400
        name = f"promoter-invokes-{date.today().isoformat()}.xlsx"
        return _xlsx_response(_invoke_workbook(rows), name)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 502


@app.route("/api/companies.xlsx", methods=["POST"])
def companies_xlsx():
    try:
        data = request.get_json(silent=True) or {}
        min_cr = to_num(data.get("min_pledge_cr")) or 0.0
        rows = data.get("companies")
        if not isinstance(rows, list) or not rows:
            rows, _meta = list_pledged_companies()
        companies = _companies_meeting_threshold(rows, min_cr)
        if not companies:
            return jsonify({"error": "No companies meet that min pledge value."}), 400
        promoters = list_promoters_for_companies(companies)
        name = f"promoter-pledges-{date.today().isoformat()}.xlsx"
        return _xlsx_response(_start_workbook(companies, promoters, min_cr), name)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 502


@app.errorhandler(500)
def _server_error(_e):
    return jsonify({"error": "Server error while opening this company."}), 500


@app.route("/api/company", methods=["POST"])
def company():
    try:
        data = request.get_json(force=True) or {}
        name = (data.get("name") or "").strip()
        symbol = (data.get("symbol") or "").strip()
        scripcode = (data.get("scripcode") or "").strip()
        isin = (data.get("isin") or "").strip()
        if not name:
            return jsonify({"error": "Choose a company from the list."}), 400
        result = load_company(
            name,
            symbol=symbol or None,
            scripcode=scripcode or None,
            isin=isin or None,
        )
        return jsonify(result)
    except Exception as e:
        traceback.print_exc()
        return jsonify({"error": str(e)}), 502


if __name__ == "__main__":
    port = int(os.environ.get("PORT", "5000"))
    print(f"Promoter pledges -> http://0.0.0.0:{port}")
    app.run(host="0.0.0.0", port=port, threaded=True, debug=False)
